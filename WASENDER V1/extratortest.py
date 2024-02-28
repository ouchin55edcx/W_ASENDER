from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

def initialize_driver(driver_path):
    options = Options()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    s = Service(driver_path)
    driver = webdriver.Chrome(service=s, options=options)
    return driver

def create_excel_sheet(sheet_name):
    wb = Workbook()
    ws = wb.active
    headlines = ['name', 'phone', 'country', 'state', 'city', 'zip_code']
    ws.append(headlines)
    wb.save(sheet_name)

def write_to_excel(data, sheet_name):
    wb = load_workbook(sheet_name)
    ws = wb.active
    ws.append(data)
    wb.save(sheet_name)

def main(country, category):
    sheet_name = f'{country}_{category}_data.xlsx'
    create_excel_sheet(sheet_name)
    chrome_driver_path = r'C:\Users\fettah-cach\Desktop\chromedriver-win64\chromedriver.exe'
    driver = initialize_driver(chrome_driver_path)
    
    search_url = f'https://www.google.com/maps/search/{category}+{country}'

    try:
        driver.get(search_url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//h1')))
    except Exception as e:
        print('Error loading the search page:', e)
        driver.quit()
        return

    while True:
        try:
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            results = soup.find_all('div', class_='section-result')
            
            for result in results:
                try:
                    name = result.find('h3', class_='section-result-title').text.strip()
                except:
                    name = 'N/A'

                try:
                    phone_element = result.find_element(By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]')
                    phone = phone_element.text.strip() if phone_element else 'N/A'
                except:
                    phone = 'N/A'

                # Click on the result to get details
                result.click()
                
                # Wait for the details page to load
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//div[@data-section-id="ap"]')))
                
                # Extract country, state, city, and zip code
                try:
                    address_element = driver.find_element(By.XPATH, '//div[@data-section-id="ap"]')
                    address_text = address_element.text
                    
                    # Implement your own parsing logic here to extract country, state, city, and zip code
                    # For example, split the address_text based on commas and spaces
                    
                    # Sample parsing logic, modify as needed
                    address_parts = address_text.split(', ')
                    country = address_parts[-1]
                    state = address_parts[-2]
                    city = address_parts[-3]
                    zip_code = address_parts[-4]
                    
                except Exception as e:
                    print('Error extracting address:', e)
                    country = state = city = zip_code = 'N/A'
                
                data = [name, phone, country, state, city, zip_code]
                write_to_excel(data, sheet_name)
                
                # Go back to search results
                driver.back()

            # Check if there's a "Next" button and click it
            next_button = driver.find_element(By.ID, "n7lv7yjyC35__section-pagination-button-next")
            next_button.click()

        except Exception as e:
            print('An error occurred:', e)
            break  # Break the loop if an error occurs or there are no more pages

    driver.quit()

if __name__ == "__main__":
    country = input("Enter the country: ")
    category = input("Enter the category (e.g., doctor): ")
    main(country, category)
